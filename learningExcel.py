# Learning to read  and write excel files

import openpyxl, os
os.chdir('C:\\Users\\rjashari\\Documents')


workbook = openpyxl.load_workbook('example.xlsx') # loads example.xlsx workbook from directory specified above
sheet = workbook.get_sheet_by_name('Sheet') # if you know the name of the sheet, you can call it 
workbook.get_sheet_names() # if you don't know the name of the sheet, this line will return a string value with all the sheet names.

cell = sheet['A1'] # returns a cell object into variable cell
print(str(cell.value)) # returns the string value representation of A1 cell
# print(str(sheet['A1'].value) # will return the same thing
print(sheet.cell(row=1, column=2)) # row and columns begin at 1 not 0 like python usually does.
# this will return B1

    
workbook2 = openpyxl.load_workbook('example4.xlsx')
sheet2 = workbook.get_sheet_by_name('Sheet1')
for i in range(1,12): # in mathematical terms this is [1,8) which means include 1 but not 8. So only 1-7 will show
    print(i, sheet.cell(row=i, column=2).value) # displays rows 1 through 7 in the command line


# Learning to write excel files

import openpyxl, os

wb = openpyxl.Workbook() # creating a new blank excel file
sheet = wb.get_sheet_by_name('Sheet') # add this as a variable instead of writing it everytime.
sheet['A1'] = 42 # assign 42 to A1 and hello to a2
sheet['A2'] = 'This sheet was automatically created using Python. Dope!'

os.chdir('C:\\Users\\rjashari\\Documents')
''' so far this file is stored into python. It is not an excel file yet.
So we change the directory and save it using the codoe below. '''

wb.save('example.xlsx') # saves it as example.xlsx into users\rjashari\documents

sheet2 = wb.create_sheet() # creates a new sheet into the workbook and saves it into variable sheets2

sheet2.title = 'New Sheet Name' # change the sheet name if you don't like it

sheet2['A1'] = 'This is sheet 2 and will NOT show in example.xlxs'

print(wb.get_sheet_names()) # will return all the sheet names

wb.save('example2.xlsx') # Saves new example2.xlxs doc with A1 = 42 and A2 = hello and 2 total worksheets

sheet3 = wb.create_sheet(index=0, title='This Sheet is First') # in case you want a sheet created first

sheet3['A1'] = 'This is ''This Sheet is First'' and will ONLY show in example3.xlxs'

wb.save('example3.xlsx') # creates new excel file and saves everything from above down to the last line



